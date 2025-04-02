import os
import json
import logging
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from google_maps_scraper import GoogleMapsScraper
from config import (
    SEARCH_TERM,
    MAX_RESULTS,
    OUTPUT_DIR,
    DEBUG_MODE,
    SAVE_SCREENSHOTS,
    EXPORT_TO_EXCEL,
    EXCEL_SHEET_NAME
)

# Configure logging
log_level = logging.DEBUG if DEBUG_MODE else logging.INFO
logging.basicConfig(
    level=log_level,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(f'scraper_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
    ]
)
logger = logging.getLogger(__name__)

def ensure_output_dir():
    """Create output directory if it doesn't exist."""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        logger.info(f"Created output directory: {OUTPUT_DIR}")

def save_screenshot(driver, error_type):
    """Save a screenshot when an error occurs."""
    if SAVE_SCREENSHOTS:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        screenshot_path = os.path.join(OUTPUT_DIR, f"error_{error_type}_{timestamp}.png")
        driver.save_screenshot(screenshot_path)
        logger.info(f"Screenshot saved: {screenshot_path}")

def export_to_excel(results, output_file):
    """Export results to an Excel file with formatting."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = EXCEL_SHEET_NAME

        # Define headers
        headers = ['Name', 'Rating', 'Review Count', 'Link']
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row, place in enumerate(results, 2):
            ws.cell(row=row, column=1, value=place['name'])
            ws.cell(row=row, column=2, value=place.get('rating', 'N/A'))
            ws.cell(row=row, column=3, value=place.get('review_count', 'N/A'))
            ws.cell(row=row, column=4, value=place['link'])
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 100)
        
        # Save the workbook
        wb.save(output_file)
        logger.info(f"Results exported to Excel: {output_file}")
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {str(e)}")
        raise

def main():
    """Main function to run the Google Maps scraper."""
    scraper = None
    try:
        # Ensure output directory exists
        ensure_output_dir()
        
        # Initialize and run scraper
        scraper = GoogleMapsScraper()
        logger.info(f"Starting search for: {SEARCH_TERM}")
        results = scraper.search_places(SEARCH_TERM, MAX_RESULTS)
        
        if not results:
            logger.warning("No results found!")
            if SAVE_SCREENSHOTS:
                save_screenshot(scraper.driver, "no_results")
            return
        
        # Generate timestamp for file names
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Save results to a JSON file
        json_file = os.path.join(
            OUTPUT_DIR,
            f"google_maps_results_{SEARCH_TERM.replace(' ', '_')}_{timestamp}.json"
        )
        
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        
        logger.info(f"Found {len(results)} places")
        logger.info(f"Results saved to {json_file}")
        
        # Export to Excel if enabled
        if EXPORT_TO_EXCEL:
            excel_file = os.path.join(
                OUTPUT_DIR,
                f"google_maps_results_{SEARCH_TERM.replace(' ', '_')}_{timestamp}.xlsx"
            )
            export_to_excel(results, excel_file)
        
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        if scraper and SAVE_SCREENSHOTS:
            save_screenshot(scraper.driver, "error")
        raise
    finally:
        if scraper:
            scraper.close()

if __name__ == "__main__":
    main() 